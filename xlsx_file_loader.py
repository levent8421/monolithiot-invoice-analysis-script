from openpyxl import load_workbook


class XlsxFileLoader:
    def __init__(self, filename):
        self._file = filename
        self._wb = None
        self._st = None

    def open(self):
        print('Open xlsx file: [%s]' % self._file)
        self._wb = load_workbook(self._file, data_only=True)

    def close(self):
        if self._wb:
            print('Close xlsx file: [%s]' % self._file)
            self._wb.close()
        self._wb = None

    def open_sheet(self, name):
        self._st = self._wb[name]
        return self

    @property
    def sheets(self):
        return self._wb.sheetnames

    def __enter__(self):
        self.open()
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        self.close()

    def read_cell(self, cell_name):
        return self._st[cell_name].value

    @property
    def sheet(self):
        return self._st


def run_test():
    loader = XlsxFileLoader('source/template_3.xlsx')
    loader.open()
    sheets = loader.sheets
    print('Loaded sheets: %s' % ','.join(sheets))
    sheet = loader.open_sheet(sheets[0])
    print(sheet)
    loader.close()


if __name__ == '__main__':
    run_test()
